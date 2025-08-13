from flask import Flask, render_template, request, jsonify, session
import json
import os
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = 'home_instead_raffle_2024'

DATA_FILE = 'raffle_data.json'
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return {'employees': {}}

def save_data(data):
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def process_excel_file(filepath):
    """
    Process Excel file and extract employee names using openpyxl.
    This function will try to find employee names in common column patterns.
    """
    try:
        # Load the Excel file
        workbook = load_workbook(filepath)
        sheet = workbook.active
        
        # Get all data from the sheet
        data = []
        headers = []
        
        # Read headers (first row)
        for cell in sheet[1]:
            headers.append(str(cell.value) if cell.value else "")
        
        # Read all data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])
        
        # Common column name patterns that might contain employee names
        name_patterns = ['name', 'caregiver', 'employee', 'staff']
        
        employees = []
        name_col_index = None
        detected_name_column = None
        
        # Find the name column by checking headers
        for i, header in enumerate(headers):
            if header and any(pattern in header.lower() for pattern in name_patterns):
                name_col_index = i
                detected_name_column = header
                break
        
        if name_col_index is not None:
            # Extract names from the identified column
            for row in data:
                if len(row) > name_col_index and row[name_col_index]:
                    name = str(row[name_col_index]).strip()
                    if (len(name) > 2 and len(name) < 50 and 
                        any(c.isalpha() for c in name) and 
                        name.lower() not in ['none', 'null', 'nan', '']):
                        employees.append(name)
        else:
            # If no obvious name column, check first few columns for potential names
            for col_index in range(min(3, len(headers))):  # Check first 3 columns
                for row in data:
                    if len(row) > col_index and row[col_index]:
                        potential_name = str(row[col_index]).strip()
                        if (len(potential_name) > 2 and len(potential_name) < 50 and 
                            any(c.isalpha() for c in potential_name) and
                            potential_name.lower() not in ['none', 'null', 'nan', '']):
                            employees.append(potential_name)
        
        workbook.close()
        
        return {
            'success': True,
            'employees': list(set(employees)),  # Remove duplicates
            'total_rows': len(data),
            'columns': headers,
            'detected_name_column': detected_name_column
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@app.route('/')
def index():
    return render_template('dashboard.html')

@app.route('/api/employees', methods=['GET'])
def get_employees():
    data = load_data()
    return jsonify(data['employees'])

@app.route('/api/employee', methods=['POST'])
def add_employee():
    data = load_data()
    employee_name = request.json.get('name', '').strip()
    
    if not employee_name:
        return jsonify({'error': 'Employee name is required'}), 400
    
    if employee_name not in data['employees']:
        data['employees'][employee_name] = {
            'entries': 0,
            'activities': [],
            'created_at': datetime.now().isoformat()
        }
        save_data(data)
    
    return jsonify({'success': True, 'employee': data['employees'][employee_name]})

@app.route('/api/employee/<employee_name>/add_entry', methods=['POST'])
def add_entry(employee_name):
    data = load_data()
    activity = request.json.get('activity', '')
    entry_value = request.json.get('entries', 1)
    
    if employee_name not in data['employees']:
        return jsonify({'error': 'Employee not found'}), 404
    
    data['employees'][employee_name]['entries'] += entry_value
    data['employees'][employee_name]['activities'].append({
        'activity': activity,
        'entries': entry_value,
        'date': datetime.now().isoformat()
    })
    
    save_data(data)
    return jsonify({'success': True, 'employee': data['employees'][employee_name]})

@app.route('/api/employee/<employee_name>', methods=['DELETE'])
def delete_employee(employee_name):
    data = load_data()
    if employee_name in data['employees']:
        del data['employees'][employee_name]
        save_data(data)
        return jsonify({'success': True})
    return jsonify({'error': 'Employee not found'}), 404

@app.route('/api/employee/<employee_name>/reset_points', methods=['POST'])
def reset_employee_points(employee_name):
    data = load_data()
    if employee_name not in data['employees']:
        return jsonify({'error': 'Employee not found'}), 404
    
    data['employees'][employee_name]['entries'] = 0
    data['employees'][employee_name]['activities'].append({
        'activity': 'Points reset to 0',
        'entries': 0,
        'date': datetime.now().isoformat()
    })
    
    save_data(data)
    return jsonify({'success': True, 'employee': data['employees'][employee_name]})

@app.route('/api/reset', methods=['POST'])
def reset_data():
    data = {'employees': {}}
    save_data(data)
    return jsonify({'success': True})

@app.route('/api/import_excel', methods=['POST'])
def import_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload .xlsx or .xls files only'}), 400
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Process the Excel file
        result = process_excel_file(filepath)
        
        # Clean up uploaded file
        os.remove(filepath)
        
        if not result['success']:
            return jsonify({'error': f'Failed to process Excel file: {result["error"]}'}), 400
        
        # Load current data
        data = load_data()
        
        # Add employees from Excel file
        added_count = 0
        for employee_name in result['employees']:
            if employee_name not in data['employees']:
                data['employees'][employee_name] = {
                    'entries': 0,
                    'activities': [],
                    'created_at': datetime.now().isoformat(),
                    'imported_from_excel': True
                }
                added_count += 1
        
        # Save updated data
        save_data(data)
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {added_count} new employees from Excel file',
            'total_employees_found': len(result['employees']),
            'new_employees_added': added_count,
            'existing_employees_skipped': len(result['employees']) - added_count,
            'file_info': {
                'total_rows': result['total_rows'],
                'columns': result['columns'],
                'detected_name_column': result['detected_name_column']
            }
        })
        
    except Exception as e:
        return jsonify({'error': f'An error occurred while processing the file: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)